# -*- coding: utf-8 -*-


# like the name says this code creates a dictionary from excel file
# this module needs this openpyxl package however
from openpyxl import load_workbook

# I created two methods since one might need only a dictionary
# workbook's sheet names are useful when trying to import the data to a database.

def excel_to_dictionary(excel_file,dictionary):
    wb = load_workbook(excel_file)
    
    
    sheets_array = wb.get_sheet_names()
    for sheet_name in sheets_array:
        actual_sheet = wb.get_sheet_by_name(sheet_name)
        
        for x in range(2,(actual_sheet.max_row + 1)):
            dictionary.update({str(sheet_name)+str(x-1) :[]})
            for i in range(1,(actual_sheet.max_column + 1)):
                dictionary[str(sheet_name)+str(x-1)].append(actual_sheet.cell(row=x, column=i).value)
    


# I'll go through this since these both are practically the same function
def excel_to_dictionary(excel_file,dictionary,workbook_names):
    #loading the excel file in a variable
    wb = load_workbook(excel_file)
    
    #variable so that we can iterate all the sheets in the chosen excel file
    sheets_array = wb.get_sheet_names()
    # iterating over sheet names
    for sheet_name in sheets_array:
        
        # creating a varible which is one of the sheets in the excel file
        actual_sheet = wb.get_sheet_by_name(sheet_name)
        
        # appending the sheet name to an array, which will ease the data extraction in some module
        workbook_names.append(sheet_name)
        
        # there is 2 because we ignore the first row in every sheet.
        # First column is usually about the names of the columns
        # in first row we iterate all the rows and the second is about the columns, 
        # so a basic double for loop to move in 2D space
        for x in range(2,(actual_sheet.max_row + 1)):
            # first we create empty list in dictionary with the sheet name plus x-1 value
            dictionary.update({str(sheet_name)+str(x-1) :[]})
            # moving in column space
            for i in range(1,(actual_sheet.max_column + 1)):
                # at the end this dictionary['index'] contains all the values in one row
                dictionary[str(sheet_name)+str(x-1)].append(actual_sheet.cell(row=x, column=i).value)

# THIS METHOD DOESN'T RETURN ANY VALUES SO GLOBAL VARIABLES ARE NEEDED TO USE THIS.

